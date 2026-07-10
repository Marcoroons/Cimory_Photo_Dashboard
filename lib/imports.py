"""Import pipeline: parse, map, flag, dedup, append idempotently.

Runs synchronously inside the app when a user uploads a file. Fine for the
weekly batch files of a few thousand rows this tool handles.
"""

import io
import re
import hashlib

import pandas as pd
from dateutil import parser as dateparser

from lib import db
from lib.flags import gps_distance_km, haversine, DEFAULT_DAILY_LIMIT, DEFAULT_GPS_THRESHOLD_KM


# Hard cap on rows accepted from one file. Guards the app and the database
# against an oversized or crafted upload. Weekly batches are a few thousand rows.
MAX_IMPORT_ROWS = 100_000

# Canonical fields the mapper targets. required, then recommended, then optional.
CANONICAL_FIELDS = [
    ("mcm_id", "required", "MCM or agent identifier"),
    ("submission_date", "required", "Date the photo was submitted"),
    ("photo_url", "required", "Direct URL to the photo"),
    ("region", "recommended", "Region or area, the location separator"),
    ("center_name", "recommended", "Booth or centre name"),
    ("captured_at", "recommended", "Timestamp the photo was taken"),
    ("category", "recommended", "Category label, for example Tempat Wisata"),
    ("photo_ref", "recommended", "Source id or filename"),
    ("latitude", "optional", "Order latitude, where the photo was taken"),
    ("longitude", "optional", "Order longitude, where the photo was taken"),
    ("customer_latitude", "optional", "Customer latitude, the GPS far reference"),
    ("customer_longitude", "optional", "Customer longitude, the GPS far reference"),
    ("gps_distance", "optional", "Pre-computed distance in km, if present"),
]

REQUIRED_FIELDS = [f for f, kind, _ in CANONICAL_FIELDS if kind == "required"]


# ---------------------------------------------------------------------------
# Reading input
# ---------------------------------------------------------------------------

def _merge_hyperlinks(data: bytes, df: pd.DataFrame) -> pd.DataFrame:
    """Fill blank cells from their embedded hyperlink target.

    Some exports put a label in the cell and the real photo URL only in the
    hyperlink. Where a cell is empty but carries a hyperlink, use the target.
    Cells that already hold text (including the URL) are left untouched. Best
    effort, any failure leaves the plain values in place.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb[wb.sheetnames[0]]
        n_cols = df.shape[1]
        for cell_row in ws.iter_rows(min_row=2, max_row=min(ws.max_row, len(df) + 1)):
            r = cell_row[0].row - 2
            if r < 0 or r >= len(df):
                continue
            for c in cell_row:
                link = c.hyperlink
                if link is None or not getattr(link, "target", None):
                    continue
                ci = c.column - 1
                if ci >= n_cols:
                    continue
                cur = df.iat[r, ci]
                if cur is None or (isinstance(cur, float) and pd.isna(cur)) or str(cur).strip() == "":
                    df.iat[r, ci] = link.target
        wb.close()
    except Exception:
        pass
    return df


def read_upload(uploaded_file) -> pd.DataFrame:
    """Read a Streamlit UploadedFile (.csv or .xlsx) into a DataFrame.

    Only .xlsx and .csv are accepted. The legacy .xls path (xlrd) is not
    supported, which keeps the parsing attack surface small.
    """
    name = (uploaded_file.name or "").lower()
    data = uploaded_file.getvalue()
    if name.endswith(".xlsx"):
        df = pd.read_excel(io.BytesIO(data), engine="openpyxl")
        df = _merge_hyperlinks(data, df)
        return df
    return pd.read_csv(io.BytesIO(data))


def read_pasted(text: str) -> pd.DataFrame:
    """Read pasted tabular text. Tab separated if tabs are present, else comma."""
    sep = "\t" if "\t" in text else ","
    return pd.read_csv(io.StringIO(text), sep=sep)


def file_hash(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


# ---------------------------------------------------------------------------
# Mapping and transform
# ---------------------------------------------------------------------------

def _parse_date(value, dayfirst=True):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        return dateparser.parse(str(value), dayfirst=dayfirst).date().isoformat()
    except (ValueError, OverflowError, TypeError):
        return None


def _parse_ts(value, dayfirst=True):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        return dateparser.parse(str(value), dayfirst=dayfirst).isoformat()
    except (ValueError, OverflowError, TypeError):
        return None


# Normalised aliases for auto-mapping. Handles the Indonesian column headers
# (Nama Region, ID MCM, URL Foto Order, Tanggal Transaksi, and so on).
_ALIASES = {
    "mcm_id": ["idmcm", "mcmid", "mcm", "kodemcm", "agentid", "agent"],
    "submission_date": ["tanggaltransaksi", "tanggal", "tgl", "submissiondate", "date", "transactiondate"],
    "photo_url": ["urlfotoorder", "urlfoto", "fotourl", "photourl", "imageurl", "url", "foto", "photo", "image", "link"],
    "region": ["namaregion", "region", "area", "wilayah"],
    "center_name": ["namacenter", "namacentre", "center", "centre", "centername", "booth"],
    "captured_at": ["capturedat", "timestamp", "takenat", "waktu", "jam"],
    "category": ["category", "kategori", "type", "jenis"],
    "photo_ref": ["photoref", "orderid", "idorder", "filename", "namafile"],
    "latitude": ["orderlatitude", "latitude", "lat"],
    "longitude": ["orderlongitude", "longitude", "long", "lng", "lon"],
    "customer_latitude": ["customerlatitude", "customerlat", "latpelanggan"],
    "customer_longitude": ["customerlongitude", "customerlong", "customerlon", "longpelanggan"],
    "gps_distance": ["gpsdistance", "distance", "jarak"],
}


def _norm(name) -> str:
    return re.sub(r"[^a-z0-9]", "", str(name).lower())


def guess_mapping(source_cols) -> dict:
    """Best-guess canonical field to source column, exact match then substring.

    Each source column is used at most once so, for example, Order Latitude and
    Customer Latitude do not both grab the latitude field.
    """
    norm = {c: _norm(c) for c in source_cols}
    mapping: dict = {}
    used: set = set()
    for field, _kind, _desc in CANONICAL_FIELDS:
        aliases = _ALIASES.get(field, [])
        match = None
        for c in source_cols:
            if c not in used and norm[c] in aliases:
                match = c
                break
        if match is None:
            for c in source_cols:
                if c not in used and any(a in norm[c] for a in aliases):
                    match = c
                    break
        if match is not None:
            mapping[field] = match
            used.add(match)
    return mapping


def _num(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _text(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    s = str(value).strip()
    return s or None


def _row_hash(project_id, mcm_id, region, submission_date, photo_ref, photo_url):
    ref = photo_ref or photo_url or ""
    basis = "|".join(str(x or "") for x in [project_id, mcm_id, region, submission_date, ref])
    return hashlib.sha256(basis.encode("utf-8")).hexdigest()


def build_rows(df: pd.DataFrame, mapping: dict, project_id: str, config: dict,
               keep_extras: bool, existing_counts: dict, existing_urls: set,
               dayfirst: bool = True) -> tuple[list, dict]:
    """Turn a mapped DataFrame into submission dicts with flags and dedup.

    Returns (rows, summary) where summary counts new-in-file flags. Duplicate
    and idempotency handling against the database happen here too. dayfirst
    parses DD/MM/YYYY dates, which is what the Indonesian source files use.
    """
    daily_limit = config.get("daily_limit", DEFAULT_DAILY_LIMIT)
    gps_threshold = config.get("gps_threshold_km", DEFAULT_GPS_THRESHOLD_KM)
    mcm_ref = config.get("mcm_reference", {})

    def col(field):
        src = mapping.get(field)
        if src and src in df.columns:
            return df[src]
        return None

    mapped_sources = {src for src in mapping.values() if src}

    # First pass: build canonical records.
    records = []
    for i in range(len(df)):
        rec = {}
        for field, _, _ in CANONICAL_FIELDS:
            series = col(field)
            rec[field] = series.iloc[i] if series is not None else None
        rec["submission_date"] = _parse_date(rec.get("submission_date"), dayfirst)
        rec["captured_at"] = _parse_ts(rec.get("captured_at"), dayfirst)
        rec["latitude"] = _num(rec.get("latitude"))
        rec["longitude"] = _num(rec.get("longitude"))
        rec["customer_latitude"] = _num(rec.get("customer_latitude"))
        rec["customer_longitude"] = _num(rec.get("customer_longitude"))
        rec["gps_distance"] = _num(rec.get("gps_distance"))
        for field in ("mcm_id", "region", "center_name", "photo_url", "photo_ref", "category"):
            rec[field] = _text(rec.get(field))
        extras = {}
        if keep_extras:
            for c in df.columns:
                if c not in mapped_sources:
                    extras[str(c)] = _text(df[c].iloc[i])
        rec["_extras"] = extras
        records.append(rec)

    # File-level daily group totals.
    file_groups: dict = {}
    for rec in records:
        key = (rec.get("mcm_id"), rec.get("submission_date"))
        file_groups[key] = file_groups.get(key, 0) + 1

    rows = []
    seen_urls = set()
    summary = {"over_limit": 0, "duplicates": 0, "no_gps": 0, "gps_far": 0}

    for rec in records:
        mcm = rec.get("mcm_id")
        date = rec.get("submission_date")
        url = rec.get("photo_url")

        existing = existing_counts.get((mcm, str(date)), 0)
        daily_count = existing + file_groups.get((mcm, date), 0)
        over_limit = daily_count > daily_limit

        no_gps = rec.get("latitude") is None or rec.get("longitude") is None
        # Prefer the per-row customer coordinate as the reference (photo taken
        # far from the customer), else a per-MCM reference from project config.
        if rec.get("customer_latitude") is not None and rec.get("customer_longitude") is not None:
            ref_coord = (rec["customer_latitude"], rec["customer_longitude"])
        else:
            ref_coord = mcm_ref.get(mcm) if mcm else None
        dist, gps_far = gps_distance_km(
            rec.get("latitude"), rec.get("longitude"), rec.get("gps_distance"),
            ref_coord, gps_threshold,
        )

        is_dup = False
        if url:
            if url in existing_urls or url in seen_urls:
                is_dup = True
            else:
                seen_urls.add(url)

        flags = {"no_gps": no_gps, "daily_count": daily_count, "over_limit": over_limit}
        if gps_far is not None:
            flags["gps_far"] = gps_far
        if dist is not None:
            flags["gps_distance_km"] = round(dist, 2)

        row = {
            "project_id": project_id,
            "region": rec.get("region"),
            "mcm_id": mcm,
            "center_name": rec.get("center_name"),
            "submission_date": date,
            "captured_at": rec.get("captured_at"),
            "photo_url": url,
            "photo_ref": rec.get("photo_ref"),
            "latitude": rec.get("latitude"),
            "longitude": rec.get("longitude"),
            "category": rec.get("category"),
            "flags": flags,
            "is_duplicate": is_dup,
            "row_hash": _row_hash(project_id, mcm, rec.get("region"), date,
                                  rec.get("photo_ref"), url),
            "metadata": rec.get("_extras") or {},
        }
        rows.append(row)

        if over_limit:
            summary["over_limit"] += 1
        if is_dup:
            summary["duplicates"] += 1
        if no_gps:
            summary["no_gps"] += 1
        if gps_far:
            summary["gps_far"] += 1

    return rows, summary


def run_import(project_id, mapping, df, raw_bytes, filename, config, keep_extras,
               user_id, dayfirst=True) -> dict:
    """Full import: batch record, idempotent append, flags, dedup, activity log.

    Returns a summary dict for the page to display.
    """
    if len(df) > MAX_IMPORT_ROWS:
        raise ValueError(
            f"File has {len(df):,} rows, above the {MAX_IMPORT_ROWS:,} limit. "
            "Please split it into smaller files."
        )

    fhash = file_hash(raw_bytes)
    existing_counts = db.existing_daily_counts(project_id)
    existing_urls = db.existing_photo_urls(project_id)

    rows, flag_summary = build_rows(
        df, mapping, project_id, config, keep_extras, existing_counts,
        existing_urls, dayfirst=dayfirst,
    )

    batch = db.get_batch_by_hash(project_id, fhash)
    if batch is None:
        batch = db.create_batch(project_id, filename, fhash, len(rows), user_id)
    batch_id = batch["id"]
    for r in rows:
        r["batch_id"] = batch_id

    inserted = db.insert_submissions(rows)
    skipped = len(rows) - inserted
    db.update_batch_counts(batch_id, inserted, skipped)

    db.log_activity(
        project_id, user_id, "imported",
        details={
            "filename": filename,
            "rows": len(rows),
            "inserted": inserted,
            "skipped": skipped,
            "over_limit": flag_summary["over_limit"],
            "duplicates": flag_summary["duplicates"],
        },
    )
    db.invalidate()

    return {
        "total": len(rows),
        "inserted": inserted,
        "skipped": skipped,
        **flag_summary,
    }


# ---------------------------------------------------------------------------
# Mapping templates
# ---------------------------------------------------------------------------

def list_templates(project_id: str) -> list:
    client = db.get_client()
    return (
        client.table("import_templates")
        .select("*")
        .eq("project_id", project_id)
        .order("created_at", desc=True)
        .execute()
        .data
        or []
    )


def save_template(project_id: str, name: str, mapping: dict, created_by: str) -> None:
    client = db.get_client()
    client.table("import_templates").insert(
        {"project_id": project_id, "name": name, "mapping": mapping, "created_by": created_by}
    ).execute()
